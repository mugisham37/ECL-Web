"""
ECL Test Template Generator
Produces three XLSX files: PD, LGD, EAD — ready for upload to the ECL platform.
Portfolio: 110 loans across 5 segments, 6 reporting months.
"""

import random
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

OUT_DIR = Path(__file__).parent

# ─── Helpers ──────────────────────────────────────────────────────────────────

def _header_style(ws, row=1, fill_hex="1F4E79"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 30


def _freeze_and_autofit(ws, freeze="A2"):
    ws.freeze_panes = freeze
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 30)


def _add_instructions(wb, text_lines):
    # Instructions are kept in a separate README-style sheet but NOT added to the workbook
    # because the backend validates every sheet as a data sheet.
    # Documentation is in the README instead.
    pass


# ─── Portfolio master list ────────────────────────────────────────────────────

def _build_loans():
    """Return list of dicts with all static loan properties."""
    loans = []

    # ── Retail (30 loans, 22 customers) ──────────────────────────────────────
    ret_custs = [f"CUST-RET-{i:03d}" for i in range(1, 23)]
    ret_amounts = [round(random.uniform(50_000, 500_000), 2) for _ in range(30)]
    ret_eir = [round(random.uniform(0.12, 0.18), 4) for _ in range(30)]
    # assign customers: first 8 customers get 1 loan each, next 7 get 2 loans each, last 7 get 1 each
    ret_cust_map = (
        ret_custs[0:8]             # 8 single-loan customers
        + [ret_custs[8]] * 2       # customer 9 → 2 loans
        + [ret_custs[9]] * 2       # customer 10 → 2 loans
        + [ret_custs[10]] * 2      # customer 11 → 2 loans
        + [ret_custs[11]] * 2      # customer 12 → 2 loans
        + [ret_custs[12]] * 2      # customer 13 → 2 loans
        + [ret_custs[13]] * 2      # customer 14 → 2 loans
        + [ret_custs[14]] * 2      # customer 15 → 2 loans
        + ret_custs[15:22]         # 7 more single-loan customers
    )  # total = 8+2+2+2+2+2+2+2+7 = 29 → add one
    ret_cust_map.append(ret_custs[21])  # 30th loan to last customer (2 loans)

    for i in range(30):
        loans.append({
            "loan_id": f"LN-RET-{i+1:03d}",
            "customer_id": ret_cust_map[i],
            "segment": "Retail",
            "amount": ret_amounts[i],
            "eir": ret_eir[i],
            "freq": "MTH",
            "origination": date(2021, random.randint(1, 12), 1),
            "tenor_months": random.choice([24, 36, 48, 60]),
        })

    # ── Corporate (20 loans, 14 customers) ───────────────────────────────────
    cor_custs = [f"CUST-COR-{i:03d}" for i in range(1, 15)]
    cor_amounts = [round(random.uniform(2_000_000, 50_000_000), 2) for _ in range(20)]
    cor_eir = [round(random.uniform(0.08, 0.13), 4) for _ in range(20)]
    cor_cust_map = (
        cor_custs[0:8]
        + [cor_custs[8]] * 2
        + [cor_custs[9]] * 2
        + [cor_custs[10]] * 2
        + [cor_custs[11]] * 2
        + [cor_custs[12]] * 2
        + [cor_custs[13]] * 2
    )
    for i in range(20):
        loans.append({
            "loan_id": f"LN-COR-{i+1:03d}",
            "customer_id": cor_cust_map[i],
            "segment": "Corporate",
            "amount": cor_amounts[i],
            "eir": cor_eir[i],
            "freq": random.choice(["MTH", "MTH", "QTR"]),
            "origination": date(2020, random.randint(1, 12), 1),
            "tenor_months": random.choice([36, 48, 60, 84]),
        })

    # ── Mortgage (25 loans, 25 customers) ────────────────────────────────────
    mtg_amounts = [round(random.uniform(300_000, 8_000_000), 2) for _ in range(25)]
    mtg_eir = [round(random.uniform(0.10, 0.15), 4) for _ in range(25)]
    for i in range(25):
        loans.append({
            "loan_id": f"LN-MTG-{i+1:03d}",
            "customer_id": f"CUST-MTG-{i+1:03d}",
            "segment": "Mortgage",
            "amount": mtg_amounts[i],
            "eir": mtg_eir[i],
            "freq": "MTH",
            "origination": date(2018, random.randint(1, 12), 1),
            "tenor_months": random.choice([120, 180, 240]),
        })

    # ── SME (20 loans, 16 customers) ─────────────────────────────────────────
    sme_custs = [f"CUST-SME-{i:03d}" for i in range(1, 17)]
    sme_amounts = [round(random.uniform(200_000, 5_000_000), 2) for _ in range(20)]
    sme_eir = [round(random.uniform(0.14, 0.22), 4) for _ in range(20)]
    sme_cust_map = (
        sme_custs[0:12]
        + [sme_custs[12]] * 2
        + [sme_custs[13]] * 2
        + [sme_custs[14]] * 2
        + [sme_custs[15]] * 2
    )
    for i in range(20):
        loans.append({
            "loan_id": f"LN-SME-{i+1:03d}",
            "customer_id": sme_cust_map[i],
            "segment": "SME",
            "amount": sme_amounts[i],
            "eir": sme_eir[i],
            "freq": random.choice(["MTH", "MTH", "QTR"]),
            "origination": date(2022, random.randint(1, 12), 1),
            "tenor_months": random.choice([24, 36, 48]),
        })

    # ── Agriculture (15 loans, 12 customers) ─────────────────────────────────
    agr_custs = [f"CUST-AGR-{i:03d}" for i in range(1, 13)]
    agr_amounts = [round(random.uniform(100_000, 2_000_000), 2) for _ in range(15)]
    agr_eir = [round(random.uniform(0.11, 0.17), 4) for _ in range(15)]
    agr_cust_map = (
        agr_custs[0:9]
        + [agr_custs[9]] * 2
        + [agr_custs[10]] * 2
        + [agr_custs[11]] * 2
    )
    for i in range(15):
        loans.append({
            "loan_id": f"LN-AGR-{i+1:03d}",
            "customer_id": agr_cust_map[i],
            "segment": "Agriculture",
            "amount": agr_amounts[i],
            "eir": agr_eir[i],
            "freq": random.choice(["MTH", "QTR", "QTR"]),
            "origination": date(2022, random.randint(1, 6), 1),
            "tenor_months": random.choice([12, 18, 24, 36]),
        })

    return loans


def _assign_staging(loans):
    """
    Assign per-month staging to create a realistic IFRS 9 transition matrix.
    Every segment gets S1→S2, S2→S3, and cure (S2→S1) transitions so the
    PD engine produces non-zero marginal PDs for all stages.
    Returns dict: loan_id → list of 6 stage strings (Oct 2024 … Mar 2025).
    """
    # Random base for any loans not explicitly assigned a transition pattern
    base = []
    for _ in range(len(loans)):
        r = random.random()
        if r < 0.68:
            base.append("Stage 1")
        elif r < 0.90:
            base.append("Stage 2")
        else:
            base.append("Stage 3")

    # Organise loans by segment so each segment gets the full transition set
    from collections import defaultdict
    seg_buckets: dict[str, list[tuple[int, dict]]] = defaultdict(list)
    for idx, loan in enumerate(loans):
        seg_buckets[loan["segment"]].append((idx, loan))

    # Deterministic transition patterns per segment position (seg_idx)
    _PATTERNS: dict[int, list[str]] = {
        # Position 0 & 1: S1 migrates to S2 mid-period
        0: ["Stage 1", "Stage 1", "Stage 1", "Stage 2", "Stage 2", "Stage 2"],
        1: ["Stage 1", "Stage 1", "Stage 2", "Stage 2", "Stage 2", "Stage 2"],
        # Position 2 & 3: S2 deteriorates to S3
        2: ["Stage 2", "Stage 2", "Stage 2", "Stage 2", "Stage 3", "Stage 3"],
        3: ["Stage 2", "Stage 3", "Stage 3", "Stage 3", "Stage 3", "Stage 3"],
        # Position 4: cure — S2 improves back to S1
        4: ["Stage 2", "Stage 2", "Stage 2", "Stage 2", "Stage 1", "Stage 1"],
        # Position 5: deep default — S3 throughout
        5: ["Stage 3", "Stage 3", "Stage 3", "Stage 3", "Stage 3", "Stage 3"],
    }

    staging_history: dict[str, list[str]] = {}
    for seg_loans in seg_buckets.values():
        for seg_idx, (global_idx, loan) in enumerate(seg_loans):
            if seg_idx in _PATTERNS:
                staging_history[loan["loan_id"]] = _PATTERNS[seg_idx]
            else:
                staging_history[loan["loan_id"]] = [base[global_idx]] * 6

    return staging_history


# ─── PD File ──────────────────────────────────────────────────────────────────

def _last_day(y, m):
    if m == 12:
        return date(y, 12, 31)
    return date(y, m + 1, 1) - timedelta(days=1)

REPORTING_MONTHS = [
    date(2024, 10, 1),
    date(2024, 11, 1),
    date(2024, 12, 1),
    date(2025,  1, 1),
    date(2025,  2, 1),
    date(2025,  3, 1),
]


def build_pd(loans, staging_history):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PD Data"

    headers = ["Loan ID", "SEGMENT", "Reporting Month", "Staging", "Loan Amount"]
    ws.append(headers)

    row_fill_s1 = PatternFill("solid", fgColor="EAF3FB")
    row_fill_s2 = PatternFill("solid", fgColor="FFF8E1")
    row_fill_s3 = PatternFill("solid", fgColor="FDECEA")
    stage_fills = {"Stage 1": row_fill_s1, "Stage 2": row_fill_s2, "Stage 3": row_fill_s3}

    for m_idx, rep_month in enumerate(REPORTING_MONTHS):
        for loan in loans:
            stage = staging_history[loan["loan_id"]][m_idx]
            # Slight amortisation per month: 0.5% reduction per month from base amount
            amount = round(loan["amount"] * (1 - 0.005 * m_idx), 2)
            row = [loan["loan_id"], loan["segment"], rep_month, stage, amount]
            ws.append(row)
            fill = stage_fills[stage]
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Format date and amount cells
            ws.cell(ws.max_row, 3).number_format = "YYYY-MM-DD"
            ws.cell(ws.max_row, 5).number_format = '#,##0.00'

    _header_style(ws)
    _freeze_and_autofit(ws)

    _add_instructions(wb, [
        "PD Template — Probability of Default Input File",
        "",
        "REQUIRED COLUMNS:",
        "  • Loan ID         — Unique loan identifier (no duplicates per Reporting Month)",
        "  • SEGMENT         — Must exactly match a configured segment in the ECL platform",
        "  • Reporting Month — Last calendar day of the observation month (YYYY-MM-DD)",
        "  • Staging         — One of: Stage 1 | Stage 2 | Stage 3",
        "  • Loan Amount     — Outstanding balance at that reporting month (≥ 0)",
        "",
        "STAGING COLOUR GUIDE:",
        "  Blue  = Stage 1 (Performing)        Orange = Stage 2 (Under-Performing)",
        "  Red   = Stage 3 (Default/Non-Performing)",
        "",
        "TRANSITION MATRIX COVERAGE:",
        "  This file includes 110 loans × 6 months = 660 rows.",
        "  Stage transitions are embedded to produce a rich transition matrix:",
        "    - 10 loans migrate Stage 1 → Stage 2 mid-period",
        "    - 5 loans deteriorate Stage 2 → Stage 3",
        "    - 3 loans cure Stage 2 → Stage 1",
        "    - Remaining loans are stable",
        "",
        "SEGMENTS COVERED:  Retail | Corporate | Mortgage | SME | Agriculture",
        "",
        "NOTE: Loan amounts are in local currency units (LCU). No currency symbol required.",
    ])

    path = OUT_DIR / "PD_Template_2024_2025.xlsx"
    wb.save(path)
    print(f"[OK] PD  → {path}  ({sum(1 for _ in ws.iter_rows(min_row=2))} data rows)")


# ─── LGD File ─────────────────────────────────────────────────────────────────

def _collateral_for(loan, base_amount):
    seg = loan["segment"]
    re_val = mv_val = cd_val = cg_val = 0.0

    if seg == "Retail":
        # Mostly motor vehicle + cash deposit; some with nothing
        if random.random() > 0.15:
            mv_val = round(base_amount * random.uniform(0.20, 0.60), 2)
            cd_val = round(base_amount * random.uniform(0.05, 0.20), 2)

    elif seg == "Corporate":
        # Mostly real estate + corporate guarantee
        if random.random() > 0.10:
            re_val = round(base_amount * random.uniform(0.30, 0.80), 2)
            cg_val = round(base_amount * random.uniform(0.10, 0.30), 2)

    elif seg == "Mortgage":
        # Primarily real estate (property as collateral)
        if random.random() > 0.05:
            re_val = round(base_amount * random.uniform(0.70, 1.20), 2)

    elif seg == "SME":
        # Mix of real estate + motor vehicle
        if random.random() > 0.15:
            re_val = round(base_amount * random.uniform(0.20, 0.60), 2)
            mv_val = round(base_amount * random.uniform(0.10, 0.30), 2)

    elif seg == "Agriculture":
        # Corporate guarantee + motor vehicle (equipment)
        if random.random() > 0.15:
            cg_val = round(base_amount * random.uniform(0.15, 0.50), 2)
            mv_val = round(base_amount * random.uniform(0.10, 0.35), 2)

    return re_val, mv_val, cd_val, cg_val


def build_lgd(loans):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LGD Data"

    headers = [
        "Customer ID", "Loan ID", "Outstanding Amount",
        "Effective Interest Rate (EIR)",
        "Real Estate", "Motor Vehicle", "Cash Deposit", "Corporate Guarantee",
    ]
    ws.append(headers)

    march_date = date(2025, 3, 31)
    months_elapsed = 5  # Oct-24 → Mar-25

    for loan in loans:
        outstanding = round(loan["amount"] * (1 - 0.005 * months_elapsed), 2)
        re_v, mv_v, cd_v, cg_v = _collateral_for(loan, outstanding)
        row = [
            loan["customer_id"],
            loan["loan_id"],
            outstanding,
            loan["eir"],
            re_v,
            mv_v,
            cd_v,
            cg_v,
        ]
        ws.append(row)
        r = ws.max_row
        ws.cell(r, 3).number_format = '#,##0.00'
        ws.cell(r, 4).number_format = '0.0000'
        for col in (5, 6, 7, 8):
            ws.cell(r, col).number_format = '#,##0.00'
        # Shade rows with zero total collateral
        total_coll = re_v + mv_v + cd_v + cg_v
        if total_coll == 0:
            for cell in ws[r]:
                cell.fill = PatternFill("solid", fgColor="F0F0F0")

    _header_style(ws)
    _freeze_and_autofit(ws)

    # NOTE: Collateral Config reference sheet removed — the backend validates every
    # sheet as LGD data, so helper sheets must not be included in uploaded files.

    _add_instructions(wb, [
        "LGD Template — Loss Given Default Input File",
        "",
        "REQUIRED COLUMNS:",
        "  • Customer ID                   — Customer identifier; multiple loans may share one customer",
        "  • Loan ID                       — Unique loan identifier (must match EAD file exactly)",
        "  • Outstanding Amount            — Current outstanding balance at reporting date (≥ 0)",
        "  • Effective Interest Rate (EIR) — Annual rate as decimal (e.g., 0.1410 = 14.10%); must be 0–1",
        "  • Real Estate                   — Market value of real estate collateral pledged (0 if none)",
        "  • Motor Vehicle                 — Market value of vehicle/equipment collateral (0 if none)",
        "  • Cash Deposit                  — Market value of pledged cash deposit (0 if none)",
        "  • Corporate Guarantee           — Face value of corporate guarantee (0 if none)",
        "",
        "COLLATERAL CONFIG:",
        "  The platform applies haircuts and time-to-realization discounts (see 'Collateral Config' sheet).",
        "  Column names in this file MUST exactly match the collateral type names configured in the platform.",
        "",
        "PROPORTIONAL COLLATERAL:",
        "  For customers with multiple loans, the platform distributes shared collateral proportionally",
        "  by outstanding amount. Ensure each customer's collateral is entered on each of their loans.",
        "",
        "GREY ROWS = zero total collateral (LGW = 1.0 after calculation — tests the uncollateralised path).",
        "",
        "PORTFOLIO COVERAGE:  110 loans across Retail | Corporate | Mortgage | SME | Agriculture",
    ])

    path = OUT_DIR / "LGD_Template_2025_03.xlsx"
    wb.save(path)
    print(f"[OK] LGD → {path}  ({sum(1 for _ in ws.iter_rows(min_row=2))} data rows)")


# ─── EAD File ─────────────────────────────────────────────────────────────────

def _add_months(d, months):
    m = d.month - 1 + months
    y = d.year + m // 12
    mo = m % 12 + 1
    day = min(d.day, [31,28,31,30,31,30,31,31,30,31,30,31][mo-1])
    return date(y, mo, day)


def build_ead(loans, staging_history):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EAD Data"

    headers = [
        "Loan ID", "Customer ID", "SEGMENT",
        "Reporting Date", "Maturity Date", "Adjusted Maturity Date", "First Payment Date",
        "Outstanding Amount", "Repayment Frequency", "Staging (Stage)",
        "Effective Interest Rate (EIR)",
    ]
    ws.append(headers)

    reporting_date = date(2025, 3, 31)
    months_elapsed = 5

    for i, loan in enumerate(loans):
        stage_mar = staging_history[loan["loan_id"]][5]  # March 2025 staging
        outstanding = round(loan["amount"] * (1 - 0.005 * months_elapsed), 2)

        # Compute maturity date from origination + tenor
        maturity = _add_months(loan["origination"], loan["tenor_months"])
        # Ensure maturity >= reporting date (clamp forward if needed)
        if maturity <= reporting_date:
            maturity = _add_months(reporting_date, random.randint(3, 12))

        # Adjusted maturity: same for most; extended for ~10 loans
        if i in (5, 15, 25, 35, 45, 55, 65, 75, 85, 95):
            adj_maturity = _add_months(maturity, random.randint(1, 6))
        else:
            adj_maturity = maturity

        # First payment date: 1 month after origination (always ≤ reporting date)
        first_payment = _add_months(loan["origination"], 1)
        if first_payment > reporting_date:
            first_payment = _add_months(reporting_date, -6)

        # Edge case EC-07: 2 Stage 2 loans with maturity = reporting date
        if i in (8, 18) and stage_mar == "Stage 2":
            maturity = reporting_date
            adj_maturity = reporting_date

        row = [
            loan["loan_id"],
            loan["customer_id"],
            loan["segment"],
            reporting_date,
            maturity,
            adj_maturity,
            first_payment,
            outstanding,
            loan["freq"],
            stage_mar,
            loan["eir"],
        ]
        ws.append(row)
        r = ws.max_row
        for col in (4, 5, 6, 7):
            ws.cell(r, col).number_format = "YYYY-MM-DD"
        ws.cell(r, 8).number_format = "#,##0.00"
        ws.cell(r, 11).number_format = "0.0000"

        # Row colour by stage
        colours = {"Stage 1": "EAF3FB", "Stage 2": "FFF8E1", "Stage 3": "FDECEA"}
        fill = PatternFill("solid", fgColor=colours.get(stage_mar, "FFFFFF"))
        for cell in ws[r]:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    _header_style(ws)
    _freeze_and_autofit(ws)

    _add_instructions(wb, [
        "EAD Template — Exposure at Default Input File",
        "",
        "REQUIRED COLUMNS:",
        "  • Loan ID                       — Unique; must match LGD file exactly (EC-06 warning if missing in LGD)",
        "  • Customer ID                   — Links to LGD data",
        "  • SEGMENT                       — Must match PD data segments (EC-08 if missing)",
        "  • Reporting Date                — Date of this snapshot (YYYY-MM-DD); must be ≤ Maturity Date (EC-03)",
        "  • Maturity Date                 — Scheduled loan end date (≥ Reporting Date)",
        "  • Adjusted Maturity Date        — Revised scheduled end (≥ First Payment Date)",
        "  • First Payment Date            — Date of first repayment (≤ Reporting Date) (EC-04)",
        "  • Outstanding Amount            — Balance at Reporting Date (≥ 0); should match LGD file",
        "  • Repayment Frequency           — 'MTH' (monthly) or 'QTR' (quarterly) only (EC-09)",
        "  • Staging (Stage)               — Stage 1 | Stage 2 | Stage 3  (matches March 2025 in PD file)",
        "  • Effective Interest Rate (EIR) — Annual rate as decimal 0–1 (EC-05); matches LGD file",
        "",
        "EDGE CASES INCLUDED:",
        "  • 2 Stage 2 loans with Maturity Date = Reporting Date (EC-07: 1 snapshot only)",
        "  • All Stage 3 loans produce 1 snapshot only",
        "  • All 5 segments present → PD data exists for all (EC-08 satisfied)",
        "  • 15% QTR frequency loans (Agriculture + Corporate)",
        "",
        "COLOUR GUIDE:  Blue=Stage 1  |  Orange=Stage 2  |  Red=Stage 3",
        "",
        "Reporting Date for all rows: 2025-03-31",
    ])

    path = OUT_DIR / "EAD_Template_2025_03.xlsx"
    wb.save(path)
    print(f"[OK] EAD → {path}  ({sum(1 for _ in ws.iter_rows(min_row=2))} data rows)")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Building ECL test templates…")
    loans = _build_loans()
    staging = _assign_staging(loans)

    build_pd(loans, staging)
    build_lgd(loans)
    build_ead(loans, staging)

    print("\nDone. Files saved to:", OUT_DIR)
    print("\nSegment summary:")
    seg_count = {}
    for l in loans:
        seg_count[l["segment"]] = seg_count.get(l["segment"], 0) + 1
    for s, c in sorted(seg_count.items()):
        print(f"  {s:12s}: {c:3d} loans")
    print(f"  {'TOTAL':12s}: {len(loans):3d} loans")


if __name__ == "__main__":
    main()
